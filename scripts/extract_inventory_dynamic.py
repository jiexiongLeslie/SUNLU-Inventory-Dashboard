#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Extract SUNLU inventory data by matching stable Excel header names.

Column positions in the source workbook move as new daily sales/shipment
columns are added, so this script never relies on fixed column numbers.
"""

from __future__ import annotations

import argparse
import json
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


@dataclass(frozen=True)
class ColumnRule:
    rows: tuple[int, ...]
    exact: str | None = None
    includes: tuple[str, ...] = ()
    any_includes: tuple[str, ...] = ()
    excludes: tuple[str, ...] = ()


@dataclass(frozen=True)
class SheetConfig:
    sheet_pattern: str
    region: str
    header_rows: tuple[int, ...]
    data_start_row: int
    category: ColumnRule
    color: ColumnRule
    store_sku: ColumnRule
    stock: ColumnRule
    in_transit: ColumnRule | None = None


SHEET_CONFIGS: tuple[SheetConfig, ...] = (
    SheetConfig("英国站", "英国", (1, 2, 8), 9, ColumnRule((8,), includes=("品类",)), ColumnRule((8,), exact="颜色"), ColumnRule((8,), any_includes=("店铺SKU", "Store SKU")), ColumnRule((1, 2), exact="总库存"), ColumnRule((1, 2), includes=("在途", "库存"))),
    SheetConfig("全球-欧洲", "欧洲", (1, 2, 8), 9, ColumnRule((8,), includes=("品类",)), ColumnRule((8,), exact="颜色"), ColumnRule((8,), any_includes=("店铺SKU", "Store SKU")), ColumnRule((1, 2), includes=("总库存", "欧洲所有通用")), ColumnRule((1, 2), includes=("在途库存",))),
    SheetConfig("全球-美国", "美国", (1, 2, 8), 9, ColumnRule((8,), includes=("品类",)), ColumnRule((8,), exact="颜色"), ColumnRule((8,), any_includes=("店铺SKU", "Store SKU")), ColumnRule((1, 2), includes=("独立站", "库存")), ColumnRule((1, 2), includes=("在途库存",))),
    SheetConfig("德国站", "德国", (1, 7), 8, ColumnRule((7,), includes=("品类",)), ColumnRule((7,), exact="颜色"), ColumnRule((7,), any_includes=("店铺SKU", "Store SKU")), ColumnRule((1, 2), includes=("库存总计",))),
    SheetConfig("法国站", "法国", (1, 7), 8, ColumnRule((7,), includes=("品类",)), ColumnRule((7,), exact="颜色"), ColumnRule((7,), any_includes=("店铺SKU", "Store SKU")), ColumnRule((1, 2), includes=("库存总计",))),
    SheetConfig("意大利站", "意大利", (1, 7), 8, ColumnRule((7,), includes=("品类",)), ColumnRule((7,), exact="颜色"), ColumnRule((7,), any_includes=("店铺SKU", "Store SKU")), ColumnRule((1, 2), includes=("库存总计",))),
    SheetConfig("全球-加拿大", "加拿大", (1, 2, 8), 9, ColumnRule((8,), includes=("品类",)), ColumnRule((8,), exact="颜色"), ColumnRule((8,), any_includes=("店铺SKU", "Store SKU")), ColumnRule((1, 2), exact="总库存"), ColumnRule((1, 2), includes=("在途", "库存"))),
    SheetConfig("全球-澳洲", "澳洲", (1, 2, 8), 9, ColumnRule((8,), includes=("品类",)), ColumnRule((8,), exact="颜色"), ColumnRule((8,), any_includes=("店铺SKU", "Store SKU")), ColumnRule((1, 2), exact="总库存"), ColumnRule((1, 2), includes=("在途库存",))),
)

OPTIONAL_METRICS = {
    "sales_7d_avg": ColumnRule((1, 2), includes=("7天日均销量",)),
    "sales_7d": ColumnRule((1, 2), any_includes=("7天总销量", "近7天总销量")),
    "sales_14d": ColumnRule((1, 2), any_includes=("14天总销量", "近14天总销量")),
    "sales_30d": ColumnRule((1, 2), any_includes=("近30天总销量",), excludes=("TK", "全球站+TK")),
}


def normalize_text(value: Any) -> str:
    return " ".join(str(value or "").split())


def to_number(value: Any) -> float:
    if isinstance(value, (int, float)):
        return float(value)
    text = normalize_text(value).replace(",", "")
    if text in {"", "/", "-"}:
        return 0
    try:
        return float(text)
    except ValueError:
        return 0


def matches(text: str, rule: ColumnRule) -> bool:
    if not text:
        return False
    if rule.exact and text != rule.exact:
        return False
    if rule.includes and not all(keyword in text for keyword in rule.includes):
        return False
    if rule.any_includes and not any(keyword in text for keyword in rule.any_includes):
        return False
    if rule.excludes and any(keyword in text for keyword in rule.excludes):
        return False
    return True


def find_column(ws, rule: ColumnRule) -> tuple[int, int, str] | None:
    for row_index in rule.rows:
        for col_index in range(1, ws.max_column + 1):
            title = normalize_text(ws.cell(row_index, col_index).value)
            if matches(title, rule):
                return row_index, col_index, title
    return None


def find_sheet(workbook, pattern: str) -> str | None:
    return next((name for name in workbook.sheetnames if pattern in name), None)


def risk_level(stock: int, sales_7d_avg: float) -> str:
    if stock <= 0:
        return "out_of_stock"
    if sales_7d_avg <= 0:
        return "slow_moving" if stock > 200 else "no_recent_sales"
    days = stock / sales_7d_avg
    if days < 7:
        return "critical"
    if days < 14:
        return "low"
    if days > 90:
        return "overstock"
    return "healthy"


def extract_sheet(workbook, cfg: SheetConfig) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    sheet_name = find_sheet(workbook, cfg.sheet_pattern)
    report: dict[str, Any] = {
        "sheet_pattern": cfg.sheet_pattern,
        "sheet_name": sheet_name,
        "region": cfg.region,
        "records": 0,
        "columns": {},
        "missing_columns": [],
    }
    if not sheet_name:
        report["missing_columns"].append("工作表")
        return [], report

    ws = workbook[sheet_name]
    column_rules = {
        "category": cfg.category,
        "color": cfg.color,
        "store_sku": cfg.store_sku,
        "stock": cfg.stock,
    }
    if cfg.in_transit:
        column_rules["in_transit"] = cfg.in_transit
    column_rules.update(OPTIONAL_METRICS)

    columns = {key: find_column(ws, rule) for key, rule in column_rules.items()}
    for key, match in columns.items():
        if match:
            row, col, title = match
            report["columns"][key] = {"row": row, "column": col, "title": title}
    for required in ("category", "color", "stock"):
        if not columns.get(required):
            report["missing_columns"].append(required)
    if report["missing_columns"]:
        return [], report

    records: list[dict[str, Any]] = []
    last_category = ""
    for row_index in range(cfg.data_start_row, ws.max_row + 1):
        category_col = columns["category"][1]
        color_col = columns["color"][1]
        raw_category = normalize_text(ws.cell(row_index, category_col).value)
        if raw_category and not raw_category.startswith("/"):
            last_category = raw_category
        if not last_category or "总计" in last_category or "合计" in last_category:
            continue

        color = normalize_text(ws.cell(row_index, color_col).value)
        sku_match = columns.get("store_sku")
        store_sku = normalize_text(ws.cell(row_index, sku_match[1]).value) if sku_match else ""
        if not color and (not store_sku or store_sku == "/"):
            continue

        stock = round(to_number(ws.cell(row_index, columns["stock"][1]).value))
        sales_7d_avg = to_number(ws.cell(row_index, columns["sales_7d_avg"][1]).value) if columns.get("sales_7d_avg") else 0
        days_of_cover = round(stock / sales_7d_avg, 1) if sales_7d_avg > 0 else None
        records.append({
            "category": last_category,
            "color": color,
            "store_sku": store_sku,
            "stock": stock,
            "region": cfg.region,
            "in_transit": round(to_number(ws.cell(row_index, columns["in_transit"][1]).value)) if columns.get("in_transit") else 0,
            "sales_7d_avg": round(sales_7d_avg, 2),
            "sales_7d": round(to_number(ws.cell(row_index, columns["sales_7d"][1]).value)) if columns.get("sales_7d") else 0,
            "sales_14d": round(to_number(ws.cell(row_index, columns["sales_14d"][1]).value)) if columns.get("sales_14d") else 0,
            "sales_30d": round(to_number(ws.cell(row_index, columns["sales_30d"][1]).value)) if columns.get("sales_30d") else 0,
            "days_of_cover": days_of_cover,
            "risk_level": risk_level(stock, sales_7d_avg),
            "source_sheet": sheet_name,
        })

    report["records"] = len(records)
    return records, report


def extract_inventory(path: Path) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    workbook = load_workbook(path, read_only=False, data_only=True)
    all_records: list[dict[str, Any]] = []
    reports: list[dict[str, Any]] = []
    for cfg in SHEET_CONFIGS:
        records, report = extract_sheet(workbook, cfg)
        all_records.extend(records)
        reports.append(report)
    return all_records, reports


def main() -> None:
    parser = argparse.ArgumentParser(description="Extract SUNLU inventory data with dynamic header matching.")
    parser.add_argument("input", type=Path, help="Source Excel workbook")
    parser.add_argument("-o", "--output", type=Path, help="JSON output path")
    parser.add_argument("--summary", action="store_true", help="Only print parse summary")
    args = parser.parse_args()

    records, reports = extract_inventory(args.input)
    for report in reports:
        name = report["sheet_name"] or report["sheet_pattern"]
        missing = ", ".join(report["missing_columns"])
        status = f"missing: {missing}" if missing else f"{report['records']} records"
        print(f"{name}: {status}")

    if args.summary:
        return
    payload = json.dumps(records, ensure_ascii=False, indent=2)
    if args.output:
        args.output.parent.mkdir(parents=True, exist_ok=True)
        args.output.write_text(payload, encoding="utf-8")
    else:
        print(payload)


if __name__ == "__main__":
    main()
